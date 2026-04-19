"""Styled Excel report generation for completed analysis jobs."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPORT_DIR = Path(__file__).resolve().parent.parent / "tmp" / "reports"
REPORT_DIR.mkdir(parents=True, exist_ok=True)

# Style presets
ACCENT_FILL = PatternFill(start_color="4f8ef7", end_color="4f8ef7", fill_type="solid")
ALT_ROW_A = PatternFill(start_color="1e2130", end_color="1e2130", fill_type="solid")
ALT_ROW_B = PatternFill(start_color="2a2d3e", end_color="2a2d3e", fill_type="solid")
BOLD_FONT = Font(bold=True, color="000000")
NORMAL_FONT = Font(color="000000")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
CENTER = Alignment(horizontal="center", vertical="center")


def _style_header(ws, row: int, max_col: int) -> None:
    """Apply accent fill + bold white font to a header row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = ACCENT_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def _bold_labels(ws, start_row: int, end_row: int) -> None:
    """Bold the label column (A) across a range of rows."""
    for r in range(start_row, end_row + 1):
        cell = ws.cell(row=r, column=1)
        if cell.value is not None:
            cell.font = BOLD_FONT


def _auto_fit(ws) -> None:
    """Set column widths to fit the widest cell value in each column."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 4


def generate_report(
    job_id: str,
    result: dict[str, Any],
    detections_log: list[dict],
) -> str:
    """Generate a styled 3-sheet Excel report for a completed job.

    Sheet 1 "Summary"        — high-level metrics with bold labels.
    Sheet 2 "Vehicle Breakdown" — class / count / percentage.
    Sheet 3 "Detection Log"  — every counting event with alternating row fill.

    Returns the path to the saved ``.xlsx`` file.
    """
    wb = Workbook()

    breakdown = result.get("vehicle_breakdown", {})
    total = result.get("total_vehicles", 0)

    # ==================================================================
    # Sheet 1 — Summary
    # ==================================================================
    ws1 = wb.active
    ws1.title = "Summary"

    rows = [
        ("Field", "Value"),
        ("Job ID", result.get("job_id", job_id)),
        ("Total Unique Vehicles", total),
        ("Processing Duration (sec)", result.get("processing_duration_sec")),
        ("Car Count", breakdown.get("car", 0)),
        ("Truck Count", breakdown.get("truck", 0)),
        ("Bus Count", breakdown.get("bus", 0)),
        ("Motorcycle Count", breakdown.get("motorcycle", 0)),
        ("Report Generated At", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")),
    ]
    for row_data in rows:
        ws1.append(list(row_data))

    _style_header(ws1, 1, 2)
    _bold_labels(ws1, 2, len(rows))

    # White font on data cells
    for r in range(2, len(rows) + 1):
        ws1.cell(row=r, column=2).font = NORMAL_FONT

    _auto_fit(ws1)

    # ==================================================================
    # Sheet 2 — Vehicle Breakdown
    # ==================================================================
    ws2 = wb.create_sheet("Vehicle Breakdown")
    ws2.append(["Class", "Count", "% of Total"])
    for cls_name, count in breakdown.items():
        pct = f"{(count / total * 100):.1f}%" if total else "0.0%"
        ws2.append([cls_name.capitalize(), count, pct])

    _style_header(ws2, 1, 3)
    for r in range(2, ws2.max_row + 1):
        for c in range(1, 4):
            ws2.cell(row=r, column=c).font = NORMAL_FONT
            ws2.cell(row=r, column=c).alignment = CENTER

    _auto_fit(ws2)

    # ==================================================================
    # Sheet 3 — Detection Log
    # ==================================================================
    ws3 = wb.create_sheet("Detection Log")
    headers = [
        "Track ID", "Class", "Confidence", "Frame Number",
        "Timestamp (sec)", "X1", "Y1", "X2", "Y2",
    ]
    ws3.append(headers)
    _style_header(ws3, 1, len(headers))

    for idx, det in enumerate(detections_log):
        row = [
            det.get("track_id"),
            det.get("class_name"),
            det.get("confidence"),
            det.get("frame_number"),
            det.get("timestamp_sec"),
            det.get("x1"),
            det.get("y1"),
            det.get("x2"),
            det.get("y2"),
        ]
        ws3.append(row)
        r = ws3.max_row
        fill = ALT_ROW_A if idx % 2 == 0 else ALT_ROW_B
        for c in range(1, len(headers) + 1):
            cell = ws3.cell(row=r, column=c)
            cell.fill = fill
            cell.font = NORMAL_FONT
            cell.alignment = CENTER

    _auto_fit(ws3)

    # ==================================================================
    # Save
    # ==================================================================
    report_path = REPORT_DIR / f"report_{job_id}.xlsx"
    wb.save(report_path)
    return str(report_path)
